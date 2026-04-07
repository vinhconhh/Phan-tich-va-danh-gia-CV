import io
import re as _re
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.config import settings
from services.classifier import classify_cv_batch
from services.extractor import extract_text
from services.infor_extractor import extract_info
from services.local_llm_reviewer import review_cv_with_local_llm
from services.matcher import rank_multiple_cvs
from services.score import score_cv

# ─── Page config ─────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Phân tích và đánh giá CV",
    page_icon="📄",
    layout="wide",
)
st.title("Phân tích, trích xuất, so khớp và đánh giá CV")

# ─── INPUT ───────────────────────────────────────────────────────────────────

col1, col2 = st.columns(2)
with col1:
    st.subheader("📄 Upload CV (nhiều file)")
    uploaded_files = st.file_uploader(
        ".", type=["pdf", "docx"],
        accept_multiple_files=True,
        label_visibility="collapsed",
    )
with col2:
    st.subheader("💼 Job Description")
    jd_text = st.text_area(
        ".", height=200,
        placeholder="Nhập mô tả công việc...",
        label_visibility="collapsed",
    )

# ─── ANALYZE ─────────────────────────────────────────────────────────────────

if st.button("🚀 Phân Tích CV", type="primary", use_container_width=True):

    if not uploaded_files:
        st.error("Vui lòng upload ít nhất 1 CV")
        st.stop()

    total_files = len(uploaded_files)
    st.info(f"Đang xử lý **{total_files}** CV...")

    # ── Bước 1: Extract text song song ──────────────────────────────────────
    status = st.status("📥 Bước 1/3 – Đọc & trích xuất văn bản...", expanded=False)
    prog1  = st.progress(0.0, text="Extracting...")

    cv_texts: dict[str, str] = {}

    def _extract_one(f):
        f.seek(0)
        return f.name, extract_text(f)

    with ThreadPoolExecutor(max_workers=min(6, total_files)) as exe:
        futures = {exe.submit(_extract_one, f): f.name for f in uploaded_files}
        done = 0
        for future in as_completed(futures):
            name, text = future.result()
            cv_texts[name] = text
            done += 1
            prog1.progress(done / total_files, text=f"Đọc CV: {done}/{total_files}")

    prog1.empty()
    status.update(label=f"✅ Bước 1/3 – Đã đọc {total_files} CV", state="complete")

    # ── Bước 2: Classify batch + extract_info + score song song ─────────────
    status2 = st.status("🔍 Bước 2/3 – Phân loại & trích xuất thông tin...", expanded=False)
    prog2   = st.progress(0.0, text="Classifying...")

    names = list(cv_texts.keys())
    texts = list(cv_texts.values())

    # Classify toàn bộ CV 1 lần (embed_batch chạy 1 lần)
    clf_list       = classify_cv_batch(texts)
    classify_results = dict(zip(names, clf_list))
    prog2.progress(0.4, text="Trích xuất thông tin...")

    # extract_info + score_cv song song (regex-based, CPU-bound nhưng nhẹ)
    extract_results: dict = {}
    score_results:   dict = {}

    def _process_one(name: str, text: str):
        info  = extract_info(text)
        score = score_cv(text, info)
        return name, info, score

    done = 0
    with ThreadPoolExecutor(max_workers=min(8, total_files)) as exe:
        futures2 = {exe.submit(_process_one, n, t): n for n, t in cv_texts.items()}
        for future in as_completed(futures2):
            name, info, score = future.result()
            extract_results[name] = info
            score_results[name]   = score
            done += 1
            prog2.progress(0.4 + 0.6 * done / total_files,
                           text=f"Xử lý: {done}/{total_files}")

    prog2.empty()
    status2.update(label="✅ Bước 2/3 – Phân loại & trích xuất xong", state="complete")

    # ── Lưu session ─────────────────────────────────────────────────────────
    st.session_state.update({
        "cv_texts":        cv_texts,
        "classify_results": classify_results,
        "extract_results": extract_results,
        "score_results":   score_results,
        "jd_text":         jd_text,
        "analyzed":        True,
    })
    st.success(f"✅ Phân tích hoàn tất {total_files} CV!", icon="🎉")

# ─── HIỂN THỊ KẾT QUẢ ────────────────────────────────────────────────────────

if st.session_state.get("analyzed", False):
    cv_texts         = st.session_state["cv_texts"]
    classify_results = st.session_state["classify_results"]
    extract_results  = st.session_state["extract_results"]
    score_results    = st.session_state["score_results"]
    saved_jd_text    = st.session_state.get("jd_text", "")

    tab1, tab2, tab3, tab4 = st.tabs([
        "🏷️ Phân loại",
        "📋 Trích xuất thông tin",
        "🔍 So khớp JD-CV",
        "⭐ Đánh giá",
    ])

    # ── TAB 1: Phân loại ─────────────────────────────────────────────────────
    with tab1:
        st.header("1. Phân Loại CV")
        for name in cv_texts:
            clf = classify_results[name]
            st.subheader(name)
            c1, c2 = st.columns(2)
            c1.metric("Danh mục",    clf["category"])
            c2.metric("Độ tin cậy", f"{clf['confidence']}%")
            if clf["all_scores"]:
                st.subheader("Chi tiết phân loại (%)")
                st.bar_chart(clf["all_scores"])
            st.divider()

    # ── TAB 2: Trích xuất ────────────────────────────────────────────────────
    with tab2:
        st.header("2. Trích Xuất Thông Tin")

        rows = []
        for name in cv_texts:
            info = extract_results[name]
            rows.append({
                "📄 File":        name,
                "👤 Họ tên":      info.get("name", "N/A"),
                "📧 Email":       info.get("email", "N/A"),
                "📞 Điện thoại":  info.get("phone", "N/A"),
                "🎓 Học vấn":     "\n".join(info.get("education", [])) or "N/A",
                "💼 Kinh nghiệm": "\n".join(info.get("experience", [])) or "N/A",
                "🛠️ Kỹ năng":    ", ".join(info.get("skills", [])) or "N/A",
            })

        df_summary = pd.DataFrame(rows)
        st.subheader("📊 Bảng tổng hợp")
        st.dataframe(
            df_summary,
            use_container_width=True,
            hide_index=True,
            column_config={
                "📄 File":        st.column_config.TextColumn(width="medium"),
                "👤 Họ tên":      st.column_config.TextColumn(width="small"),
                "📧 Email":       st.column_config.TextColumn(width="medium"),
                "📞 Điện thoại":  st.column_config.TextColumn(width="small"),
                "🎓 Học vấn":     st.column_config.TextColumn(width="large"),
                "💼 Kinh nghiệm": st.column_config.TextColumn(width="large"),
                "🛠️ Kỹ năng":    st.column_config.TextColumn(width="medium"),
            },
        )

        # ── Export Excel ──────────────────────────────────────────────────────
        df_export = pd.DataFrame([{
            "File":        r["📄 File"],
            "Họ tên":      r["👤 Họ tên"],
            "Email":       r["📧 Email"],
            "Điện thoại":  r["📞 Điện thoại"],
            "Học vấn":     r["🎓 Học vấn"],
            "Kinh nghiệm": r["💼 Kinh nghiệm"],
            "Kỹ năng":     r["🛠️ Kỹ năng"],
        } for r in rows])

        _ILLEGAL = _re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")

        def clean_for_excel(val):
            return _ILLEGAL.sub("", val) if isinstance(val, str) else val

        df_export = df_export.applymap(clean_for_excel)

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df_export.to_excel(writer, index=False, sheet_name="CV Data")
            ws = writer.sheets["CV Data"]

            header_font  = Font(bold=True, color="FFFFFF", size=11)
            header_fill  = PatternFill(start_color="1F4287", end_color="1F4287", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_align   = Alignment(horizontal="left", vertical="top", wrap_text=True)
            thin_border  = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"),  bottom=Side(style="thin"),
            )

            for cell in ws[1]:
                cell.font = header_font; cell.fill = header_fill
                cell.alignment = header_align; cell.border = thin_border

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = cell_align; cell.border = thin_border

            for col in ws.columns:
                max_len = max(
                    (max(len(l) for l in str(c.value).split("\n")) if c.value else 0)
                    for c in col
                )
                ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 4, 12), 50)

            ws.freeze_panes = "A2"

        st.download_button(
            label="⬇️ Tải xuống Excel",
            data=buf.getvalue(),
            file_name="cv_extracted.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.divider()
        st.subheader("🔎 Chi tiết từng CV")
        for name in cv_texts:
            info = extract_results[name]
            with st.expander(f"📄 {name}", expanded=False):
                df_personal = pd.DataFrame([{
                    "Họ tên":     info.get("name", "N/A"),
                    "Email":      info.get("email", "N/A"),
                    "Điện thoại": info.get("phone", "N/A"),
                }])
                st.markdown("**👤 Cá nhân**")
                st.dataframe(df_personal, use_container_width=True, hide_index=True)

                if info.get("education"):
                    st.markdown("**🎓 Học vấn**")
                    for line in info["education"]:
                        st.markdown(f"- {line}")

                if info.get("experience"):
                    st.markdown("**💼 Kinh nghiệm**")
                    for line in info["experience"]:
                        st.markdown(f"- {line}")

                if info.get("skills"):
                    st.markdown("**🛠️ Kỹ năng**")
                    cols = st.columns(4)
                    for i, skill in enumerate(info["skills"]):
                        cols[i % 4].markdown(
                            f'<span style="background:#1f4287;color:white;'
                            f'padding:3px 10px;border-radius:12px;font-size:13px">{skill}</span>',
                            unsafe_allow_html=True,
                        )

    # ── TAB 3: So khớp JD ────────────────────────────────────────────────────
    with tab3:
        st.header("3. So Khớp JD - CV")
        if not saved_jd_text.strip():
            st.warning("Vui lòng nhập Job Description để xem kết quả so khớp")
        else:
            with st.spinner("Đang tính độ phù hợp..."):
                ranking = rank_multiple_cvs(cv_texts, saved_jd_text)

            df_rank = pd.DataFrame([
                {"Hạng": i + 1, "File CV": name, "Độ phù hợp (%)": score}
                for i, (name, score) in enumerate(ranking)
            ])
            st.dataframe(
                df_rank,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Hạng": st.column_config.NumberColumn(width="small"),
                    "Độ phù hợp (%)": st.column_config.ProgressColumn(
                        min_value=0, max_value=100, format="%.1f%%"
                    ),
                },
            )

    # ── TAB 4: Đánh giá ──────────────────────────────────────────────────────
    with tab4:
        st.header("4. Đánh Giá Tổng Thể")
        weights = {"skills": 30, "experience": 30, "education": 15, "length": 10, "keywords": 15}

        score_rows = []
        for name in cv_texts:
            s   = score_results[name]
            row = {"File CV": name, "Tổng điểm": s["total_score"]}
            for k, w in weights.items():
                row[k.upper()] = f"{round(s['breakdown'][k] * w)}/{w}"
            score_rows.append(row)

        df_scores = pd.DataFrame(score_rows)
        st.dataframe(
            df_scores,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Tổng điểm": st.column_config.ProgressColumn(
                    min_value=0, max_value=100, format="%d/100"
                )
            },
        )

        st.caption("💡 Mở từng CV bên dưới để xem đánh giá chi tiết từ AI (chạy khi mở).")

        # LLM review lazy: chỉ chạy khi user mở expander
        for name in cv_texts:
            with st.expander(f"🤖 {name}", expanded=False):
                # Dùng key riêng để cache kết quả LLM review trong session
                cache_key = f"llm_review_{name}"
                if cache_key not in st.session_state:
                    with st.spinner("AI đang phân tích..."):
                        st.session_state[cache_key] = review_cv_with_local_llm(
                            cv_text=cv_texts[name],
                            score_result=score_results[name],
                            extract_result=extract_results[name],
                            classify_result=classify_results[name],
                            jd_text=saved_jd_text,
                        )
                st.markdown(st.session_state[cache_key])